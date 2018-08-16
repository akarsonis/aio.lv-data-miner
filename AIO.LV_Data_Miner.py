from bs4 import BeautifulSoup
import requests
import os
import urllib.request
import wget

import re

import openpyxl
from openpyxl import Workbook

#________________________________________________________________________

#Crawler
#AIO.LV

def data_from_aio():

    links = ['https://aio.lv/lv/skaistumkopsanai-un-veselibai-smarzas-un-pro-kosmetika/smarzas-vinai', 'https://aio.lv/lv/skaistumkopsanai-un-veselibai-smarzas-un-pro-kosmetika/smarzas-vinam', 'https://aio.lv/lv/skaistumkopsanai-un-veselibai-smarzas-un-pro-kosmetika/trimeri', 'https://aio.lv/lv/skaistumkopsanai-un-veselibai-smarzas-un-pro-kosmetika/mutes-higienai', 'https://aio.lv/lv/skaistumkopsanai-un-veselibai-smarzas-un-pro-kosmetika/skuvekli', 'https://aio.lv/lv/skaistumkopsanai-un-veselibai-smarzas-un-pro-kosmetika/matu-veidotaji-un-taisnotaji', 'https://aio.lv/lv/skaistumkopsanai-un-veselibai-smarzas-un-pro-kosmetika/feni', 'https://aio.lv/lv/preces-skaistumkopsanai-un-veselibai-smarzas-rotaslietas/elektroniskie-termometri', 'https://aio.lv/lv/skaistumkopsanai-un-veselibai-smarzas-un-pro-kosmetika/epilatori', 'https://aio.lv/lv/preces-skaistumkopsanai-un-veselibai-smarzas-rotaslietas/asinsspiediena-meritaji']
    
    max_pages = [7, 4, 6, 2, 6, 11, 8, 1, 3, 2]
    
    for (link, max_page) in zip(links, max_pages):
        
        wb = Workbook()
        ws = wb.active
        
        ws.cell(row=1, column=1).value = 'Name'
        ws.cell(row=1, column=2).value = 'Price'
        ws.cell(row=1, column=3).value = 'Product ID'
        ws.cell(row=1, column=4).value = 'Product code'
        ws.cell(row=1, column=5).value = 'EAN code'
        ws.cell(row=1, column=6).value = 'Specs'
        ws.cell(row=1, column=7).value = 'Specs2'        
        
        name_count = 2
        price_count = 2
        product_id_count = 2
        product_code_count = 2
        ean_code_count = 2
        specs_count = 2
        dictionary_specs_count = 2            
        
        start_page = 1
        
        #Directory
        
        folder_name = link.split('/', -1)[-1]
        
        home_dir = ('C:/Users/Aleksandrs/Desktop/imgs/' + folder_name)
        
        if not os.path.isdir(home_dir):
            os.makedirs(home_dir)
        
#________________________________________________________________________
        
        while start_page < max_page + 1:
            
            link_list = []     
                    
            url = link + '?page=' + str(start_page)
            htmlContent = requests.get(url)
            soup = BeautifulSoup(htmlContent.content, 'html.parser')
            htmlContent = soup.prettify()
            
#______________________________________________________________________________
            
            #Name
            
            img_name = soup.find_all('img', {'class' : 'photo'})
            
            for single_name in img_name:
                single_name = str(single_name)
                single_name = single_name.split('=', 1)[-1]
                single_name = single_name[1:]
                single_name = single_name.split('class="photo"', 1)[0]
                single_name = single_name[:-2]
                ws.cell(row=name_count, column=1).value = single_name
                name_count += 1
            
            #Price
            
            meta_price = soup.find_all('meta', {'itemprop' : 'price'})
            
            for single_price in meta_price:
                single_price = str(single_price)
                single_price = single_price.split('"', 1)[-1]
                single_price = single_price.split('"', 1)[0]
                ws.cell(row=price_count, column=2).value = float(single_price)
                price_count += 1
#_____________________________________________________________________________
                
            #IMG
            
            img_image = soup.find_all('img', {'itemprop' : 'image'})
            
            for single_image in img_image:
                single_image = str(single_image)
                
                product_id = single_image.split('data-default-src="/img/product/', 1)[-1]
                product_id = product_id.split('/', 1)[0]
                ws.cell(row=product_id_count, column=3).value = product_id
                product_id_count += 1                
            
                single_image = single_image.split('temprop="image" src="', 1)[-1]
                single_image = 'https://aio.lv' + single_image[:-3]
                
                try:
                    file_path = "C:/Users/Aleksandrs/Desktop/imgs/" + folder_name + '/' + product_id + '.jpg'
                    wget.download(single_image, file_path)
                    print('done')
                except OSError:
                    print('no image')
            
            #Link
                    
            h2_link = soup.find_all('h2', {'itemprop' : 'name'})
                    
            for single_h2_link in h2_link:
                single_h2_link = str(single_h2_link)
                single_h2_link = single_h2_link.split('a href="', 1)[-1]
                single_h2_link = single_h2_link.split('">', 1)[0]
                link_list.append(single_h2_link)
                
    #______________________________________________________________________________
    
            #Inside links
            
            for single_link_list in link_list:
                
                list_initial_specs = []
                list_name = []
                list_specification = []
                dictionary_specs = dict()
                
                product_url = 'https://aio.lv' + single_link_list
                product_htmlContent = requests.get(product_url)
                product_soup = BeautifulSoup(product_htmlContent.content, 'html.parser')
                
                #Product code
                
                b_code = product_soup.find('b', {'itemprop' : 'mpn'})
                
                try:
                    b_code = b_code.get_text()
                    b_code = re.sub(r'(^[ \t]+|[ \t]+(?=:))', '', b_code, flags=re.M)
                    b_code = "".join([s for s in b_code.splitlines(True) if s.strip("\r\n \t")])
                except AttributeError:
                    pass
    
                ws.cell(row=product_code_count, column=4).value = b_code
                product_code_count += 1
            
                #Product code
                
                b_ean = product_soup.find('b', {'itemprop' : 'ean'})
                
                try:
                    b_ean = b_ean.get_text()
                    b_ean = re.sub(r'(^[ \t]+|[ \t]+(?=:))', '', b_ean, flags=re.M)
                    b_ean = "".join([s for s in b_ean.splitlines(True) if s.strip("\r\n \t")])
                except AttributeError:
                    pass
    
                ws.cell(row=ean_code_count, column=5).value = b_ean
                ean_code_count += 1
                print(ean_code_count)
    
    #______________________________________________________________________________
    
                #Description
                div_description = product_soup.find('div', {'itemprop' : 'description'})
    
                try:
                    div_description = div_description.get_text()
                    div_description = div_description.split()
                    div_description = ' '.join(div_description)
                    ws.cell(row=specs_count, column=6).value = str(div_description)
                except AttributeError:
                    ws.cell(row=specs_count, column=6).value = ''
    
                specs_count += 1
    #______________________________________________________________________________
                #Specs Dictionary
                
                ul_features = product_soup.find_all('ul')[-1]
                
                for line in ul_features:
                    line = str(line)
                    list_initial_specs.append(line)
                
                while True:
                    try:
                        list_initial_specs.remove('\n')
                    except ValueError:
                        break
                
                for line in list_initial_specs:
                    line = line.replace('</li>', '')
                    line = line.replace('<li>', '')
                    line = line.replace('\n', '')  
                    try:
                        line = line.split('<span>')[1].split('</span>')[0]
                    except IndexError:
                        del line
                    try:
                        list_name.append(line)
                    except NameError:
                        pass
                    
                for line in list_initial_specs:
                    line = line.replace('</li>', '')
                    line = line.replace('<li>', '')
                    line = line.replace('\n', '')  
                    try:
                        line = line.split('<p>')[1].split('</p>')[0]
                    except IndexError:
                        del line
                    try:
                        list_specification.append(line)
                    except NameError:
                        pass
                
                dictionary_specs = dict(zip(list_name, list_specification))
                ws.cell(row=dictionary_specs_count, column=7).value = str(dictionary_specs)
                dictionary_specs_count += 1            
    ##______________________________________________________________________________
            
            start_page += 1
            print(start_page)
            
            try:
                wb.save('aio.lv ' + folder_name + '.xlsx')
            except PermissionError:
                wb.save('aio.lv ' + folder_name + '(1).xlsx')
                
data_from_aio()
